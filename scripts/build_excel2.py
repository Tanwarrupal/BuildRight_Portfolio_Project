import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter


sales_sample = pd.read_csv(f"{DATA}/sales_for_excel.csv").head(100)
category_summary = pd.read_csv(f"{OUT}/revenue_by_category.csv")
monthly_trend = pd.read_csv(f"{OUT}/monthly_trend_with_ma.csv")
customer_segment = pd.read_csv(f"{OUT}/customer_segment.csv")
city_perf = pd.read_csv(f"{OUT}/city_performance.csv")
top10 = pd.read_csv(f"{OUT}/top10_products.csv")
forecast = pd.read_csv(f"{OUT}/revenue_forecast_next3.csv")

wb = Workbook()
NAVY = "1F3864"
FONT = "Arial"
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor=NAVY)
normal_font = Font(name=FONT, size=10)
note_font = Font(name=FONT, size=9, italic=True, color="777777")
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

def autofit(ws, ncols, width=16):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width

def write_df(ws, df, start_row=1, number_cols=None, formats=None):
    cols = list(df.columns)
    for j, c in enumerate(cols, start=1):
        ws.cell(row=start_row, column=j, value=c)
    style_header_row(ws, start_row, len(cols))
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = normal_font
            cell.border = border
            if formats and cols[j - 1] in formats:
                cell.number_format = formats[cols[j - 1]]
    autofit(ws, len(cols), 18)
    return start_row + len(df) + 1  # next free row

# ---------------- Sales_Sample (evidence of cleaned data) ----------------
ws1 = wb.active
ws1.title = "Sales_Clean_Sample"
ws1["A1"] = "Sample of cleaned transaction-level data (500 of 9,534 rows). Full cleaning logic: /sql/analysis.sql"
ws1["A1"].font = note_font
write_df(ws1, sales_sample, start_row=2)
ws1.freeze_panes = "A3"

# ---------------- Category_Summary ----------------
ws3 = wb.create_sheet("Category_Summary")
ws3["A1"] = "Source: SQL query (revenue_by_category) — see /sql/analysis.sql section 2"
ws3["A1"].font = note_font
next_row = write_df(ws3, category_summary, start_row=2,
                     formats={"total_revenue": "#,##0", "gross_margin": "#,##0", "margin_pct": "0.0"})
n_cats = len(category_summary)

# ---------------- Monthly_Summary ----------------
ws4 = wb.create_sheet("Monthly_Summary")
ws4["A1"] = "Source: SQL query (monthly_revenue_trend), 3-month moving avg & trend computed in Python"
ws4["A1"].font = note_font
trend_export = monthly_trend[["month", "revenue", "moving_avg_3m"]].copy()
trend_export.columns = ["Month", "Revenue", "3-Month Moving Avg"]
write_df(ws4, trend_export, start_row=2, formats={"Revenue": "#,##0", "3-Month Moving Avg": "#,##0"})
n_months = len(trend_export)

# add forecast rows right after, clearly labeled
fc_start = 2 + n_months + 2
ws4.cell(row=fc_start, column=1, value="Forecast (next 3 months, linear trend)").font = Font(name=FONT, bold=True, size=10, color=NAVY)
ws4.cell(row=fc_start + 1, column=1, value="Month")
ws4.cell(row=fc_start + 1, column=2, value="Forecast Revenue")
style_header_row(ws4, fc_start + 1, 2)
for i, r in enumerate(forecast.itertuples(index=False), start=fc_start + 2):
    ws4.cell(row=i, column=1, value=r.month).font = normal_font
    c = ws4.cell(row=i, column=2, value=r.forecast_revenue)
    c.font = normal_font
    c.number_format = "#,##0"

# ---------------- Customer_Segment ----------------
ws5 = wb.create_sheet("Customer_Segment")
ws5["A1"] = "Source: SQL query (customer_segment) — see /sql/analysis.sql section 5"
ws5["A1"].font = note_font
cs = customer_segment.copy()
cs.columns = ["Customer Type", "Num Customers", "Revenue", "Avg Revenue / Customer"]
write_df(ws5, cs, start_row=2, formats={"Revenue": "#,##0", "Avg Revenue / Customer": "#,##0"})
n_segs = len(cs)

# ---------------- City_Summary ----------------
ws6 = wb.create_sheet("City_Summary")
ws6["A1"] = "Source: SQL query (city_performance) — see /sql/analysis.sql section 7"
ws6["A1"].font = note_font
cp = city_perf.copy()
cp.columns = ["City", "Revenue", "Transactions"]
write_df(ws6, cp, start_row=2, formats={"Revenue": "#,##0"})

# ---------------- Top10 Products ----------------
ws7 = wb.create_sheet("Top10_Products")
ws7["A1"] = "Source: SQL query (top10_products) — see /sql/analysis.sql section 4"
ws7["A1"].font = note_font
t10 = top10.copy()
t10.columns = ["Product Name", "Category", "Units Sold", "Revenue"]
write_df(ws7, t10, start_row=2, formats={"Revenue": "#,##0"})

# ---------------- Dashboard ----------------
wsd = wb.create_sheet("Dashboard", 0)
wsd.sheet_view.showGridLines = False
wsd["B2"] = "BuildRight Hardware & Materials — Sales & Inventory Dashboard"
wsd["B2"].font = Font(name=FONT, bold=True, size=18, color=NAVY)
wsd["B3"] = "Jan 2024 - Jun 2026  |  Self-directed analytics project  |  SQL for cleaning & aggregation, Excel for presentation"
wsd["B3"].font = Font(name=FONT, italic=True, size=10, color="666666")

kpi_labels = ["Total Revenue", "Gross Margin %", "Total Transactions", "Active Customers"]
total_revenue = category_summary["total_revenue"].sum()
total_margin = category_summary["gross_margin"].sum()
margin_pct = total_margin / total_revenue
total_txns = 9534
active_custs = customer_segment["num_customers"].sum()
kpi_values = [total_revenue, margin_pct, total_txns, active_custs]
kpi_formats = ["#,##0", "0.0%", "#,##0", "0"]

start_col = 2
for idx, (lbl, val, fmt) in enumerate(zip(kpi_labels, kpi_values, kpi_formats)):
    col = start_col + idx * 3
    cell_lbl = wsd.cell(row=5, column=col, value=lbl)
    cell_lbl.font = Font(name=FONT, size=10, color="666666")
    cell_val = wsd.cell(row=6, column=col, value=val)
    cell_val.font = Font(name=FONT, bold=True, size=16, color=NAVY)
    cell_val.number_format = fmt

# Charts referencing the small summary sheets (fast to recalc/render)
line = LineChart()
line.title = "Monthly Revenue Trend"
line.style = 2
line.y_axis.title = "Revenue (Rs)"
line.x_axis.title = "Month"
data_ref = Reference(ws4, min_col=2, max_col=3, min_row=2, max_row=2 + n_months)
cats_ref = Reference(ws4, min_col=1, min_row=3, max_row=2 + n_months)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.height = 8
line.width = 18
wsd.add_chart(line, "B9")

bar = BarChart()
bar.type = "col"
bar.title = "Revenue by Category"
bar.y_axis.title = "Revenue (Rs)"
data_ref2 = Reference(ws3, min_col=2, max_col=2, min_row=2, max_row=2 + n_cats)
cats_ref2 = Reference(ws3, min_col=1, min_row=3, max_row=2 + n_cats)
bar.add_data(data_ref2, titles_from_data=True)
bar.set_categories(cats_ref2)
bar.height = 8
bar.width = 18
wsd.add_chart(bar, "K9")

bar2 = BarChart()
bar2.type = "bar"
bar2.title = "Revenue by Customer Segment"
data_ref3 = Reference(ws5, min_col=3, max_col=3, min_row=2, max_row=2 + n_segs)
cats_ref3 = Reference(ws5, min_col=1, min_row=3, max_row=2 + n_segs)
bar2.add_data(data_ref3, titles_from_data=True)
bar2.set_categories(cats_ref3)
bar2.height = 8
bar2.width = 18
wsd.add_chart(bar2, "B26")

wsd.column_dimensions["A"].width = 3

wb.save(f"{OUT}/BuildRight_Sales_Dashboard.xlsx")
print("saved")
